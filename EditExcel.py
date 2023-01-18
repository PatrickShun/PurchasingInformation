from openpyxl import load_workbook


class runExcel():
    def __init__(self,cityName):
        self.wb = load_workbook('result.xlsx')
        self.sheet = self.wb[cityName]

    def readExcel(self, col):
        return self.sheet[col]

    def writeExcel(self,idate,col_num):
        row_init = 2
        # print('WriteTotal:', len(idate))
        for val in range(len(idate)):
            self.sheet.cell(row_init, col_num).value = idate[val].strip()
            # print(val, idate[val])
            row_init += 1
        self.wb.save('result.xlsx')

    def parseData(self,new_data):
        raw_data = self.readExcel('B')
        print(len(raw_data))


    def dataDeduplication(self):
        pass

    def run_main(self):
        # self.parseData(['123','456'])
        x = ['1','2','3','3','3']
        y = ['3','4','5']
        x = set(x)
        y = set(y)
        print(x)

if __name__ == "__main__":
    myrunning = runExcel('石家庄')
    myrunning.run_main()
